"""
Lucas Correia
LIACS | Leiden University
Einsteinweg 55 | 2333 CC Leiden | The Netherlands

Original paper DOI: 10.1137/1.9781611977653.ch77
"""

from dotenv import dotenv_values
import random
import openpyxl
from sklearn import metrics
import pandas as pd
from helper_functions import *
from utilities import detection_class
from tqdm import tqdm

# Declare constants
SEED = 1
AD_MODE = 'us'  # or 'ss'
MODEL_NAME = 'tsadis'

# Set fixed seed for random operations
random.seed(SEED)
np.random.seed(SEED)
os.environ['PYTHONHASHSEED'] = str(SEED)

# Load variables in .env file
config = dotenv_values("../.env")

# Load directory paths from .env file
data_path = config['data_path']
model_path = config['model_path']

results = []
results_best = []
for fold_idx in range(3):
    model_seed = SEED
    model_name = MODEL_NAME + '_' + AD_MODE + '_' + str(fold_idx) + '_' + str(model_seed)
    if AD_MODE == 'us':
        data_load_path = os.path.join(data_path, '2_preprocessed', 'unsupervised', 'fold_' + str(fold_idx))
    else:
        data_load_path = os.path.join(data_path, '2_preprocessed', 'semisupervised', 'fold_' + str(fold_idx))
    model_load_path = os.path.join(model_path, model_name)

    detector = detection_class.AnomalyDetector(
        model_path=model_load_path,
        window_size=256,
        sampling_rate=2,
        original_sampling_rate=10,
        calculate_delay=True,
        reverse_window_penalty=False,
    )

    # Load data
    val_list = detector.load_pickle(os.path.join(data_load_path, 'val.pkl'))
    test_list = detector.load_pickle(os.path.join(data_load_path, 'test.pkl'))

    val_detection_score_list = []
    for val_idx, val_ts in enumerate(tqdm(val_list)):
        mps = cal_tags_mps(val_ts, win=detector.window_size)
        kdps, kdps_idx = fast_find_anomalies(mps)
        detection_score = get_score(kdps, 'sum')
        val_detection_score_list.append(detection_score)

    # Process test data
    test_detection_score_list = []
    for test_idx, test_ts in enumerate(tqdm(test_list)):
        mps = cal_tags_mps(test_ts, win=detector.window_size)
        kdps, kdps_idx = fast_find_anomalies(mps)
        detection_score = get_score(kdps, 'sum')
        test_detection_score_list.append(detection_score)

    # Evaluate the model
    threshold = detector.unsupervised_threshold(val_detection_score_list)

    groundtruth_labels, predicted_labels, total_delays = detector.evaluate(
        input_list=test_list,
        detection_score_list=test_detection_score_list,
        threshold=threshold,
    )

    results.append({
        'Seed': model_seed,
        'Fold': fold_idx,
        'F1': metrics.f1_score(groundtruth_labels, predicted_labels, zero_division=0.0),
        'Precision': metrics.precision_score(groundtruth_labels, predicted_labels, zero_division=0.0),
        'Recall': metrics.recall_score(groundtruth_labels, predicted_labels, zero_division=0.0),
        'Delay': np.mean(total_delays),
        'Threshold': threshold
    })

    f1_list = []
    reduced_test_detection_score = np.concatenate(test_detection_score_list).ravel()
    percentile_array = np.arange(0, 100.01, 0.01)
    for threshold_percentile in percentile_array:
        threshold_temp = np.percentile(reduced_test_detection_score, threshold_percentile)
        groundtruth_labels, predicted_labels, _ = detector.evaluate(
            input_list=test_list,
            detection_score_list=test_detection_score_list,
            threshold=threshold_temp,
        )
        f1_list.append(metrics.f1_score(groundtruth_labels, predicted_labels, zero_division=0.0))
    f1_list = np.vstack(f1_list)
    threshold_best = np.percentile(reduced_test_detection_score, percentile_array[np.argmax(f1_list)]).astype(float)

    groundtruth_labels_best, predicted_labels_best, total_delays_best = detector.evaluate(
        input_list=test_list,
        detection_score_list=test_detection_score_list,
        threshold=threshold_best,
    )

    results_best.append({
        'Seed': model_seed,
        'Fold': fold_idx,
        'F1': metrics.f1_score(groundtruth_labels_best, predicted_labels_best, zero_division=0.0),
        'Precision': metrics.precision_score(groundtruth_labels_best, predicted_labels_best, zero_division=0.0),
        'Recall': metrics.recall_score(groundtruth_labels_best, predicted_labels_best, zero_division=0.0),
        'Delay': np.mean(total_delays_best),
        'Threshold': threshold_best
    })

results = pd.DataFrame(results)
results_best = pd.DataFrame(results_best)

if not os.path.isfile(os.path.join(model_path, 'results.xlsx')):
    # Create and save a valid Excel file
    wb = openpyxl.Workbook()
    wb.save(os.path.join(model_path, 'results.xlsx'))

# Use a try-finally block to ensure proper handling
try:
    with pd.ExcelWriter(os.path.join(model_path, 'results.xlsx'), mode='a', if_sheet_exists='overlay') as writer:
        results.to_excel(writer, index=False, sheet_name=MODEL_NAME + '_' + AD_MODE)
        results_best.to_excel(writer, index=False, sheet_name=MODEL_NAME + '_' + AD_MODE + '_best')
finally:
    # Cleanup: Remove default 'Sheet' if it exists
    try:
        workbook = openpyxl.load_workbook(os.path.join(model_path, 'results.xlsx'))
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        workbook.save(os.path.join(model_path, 'results.xlsx'))
    except Exception as e:
        print(f"Error cleaning up sheets: {e}")
