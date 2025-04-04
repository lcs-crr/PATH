"""
Lucas Correia
LIACS | Leiden University
Einsteinweg 55 | 2333 CC Leiden | The Netherlands
"""

import os

os.environ["CUDA_VISIBLE_DEVICES"] = "-1"
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'  # or any {'0', '1', '2'}
import numpy as np
from dotenv import dotenv_values
from utilities import detection_class
from sklearn import metrics
import pandas as pd
import openpyxl
import tensorflow as tf

# Declare constants
AD_MODE = 'us'  # or 'ss'
MODEL_NAME = 'tevae'  # or 'tcnae', 'omnianomaly', 'sisvae', 'lwvae', 'vsvae', 'vasp'

# Load variables in .env file
config = dotenv_values("../.env")

# Load directory paths from .env file
data_path = config['data_path']
model_path = config['model_path']

results = []
results_best = []
# Iterate over all seeds and folds
for model_seed in range(1, 4):
    for fold_idx in range(3):
        # Declare model name and paths
        model_name = MODEL_NAME + '_' + AD_MODE + '_' + str(fold_idx) + '_' + str(model_seed)
        if AD_MODE == 'us':
            data_load_path = os.path.join(data_path, '2_preprocessed', 'unsupervised', 'fold_' + str(fold_idx))
        else:
            data_load_path = os.path.join(data_path, '2_preprocessed', 'semisupervised', 'fold_' + str(fold_idx))
        model_load_path = os.path.join(model_path, model_name)

        # Load tf.data to get window_size
        tfdata_train = tf.data.Dataset.load(os.path.join(data_load_path, 'train'))

        detector = detection_class.AnomalyDetector(
            model_path=model_load_path,
            window_size=tfdata_train.element_spec.shape[0],
            sampling_rate=2,
            original_sampling_rate=10,
            calculate_delay=True,
            label_keyword='normal',
        )

        # Load data
        val_list = detector.load_pickle(os.path.join(data_load_path, 'val.pkl'))
        test_list = detector.load_pickle(os.path.join(data_load_path, 'test.pkl'))

        # Load inference results for validation data
        val_detection_score_list = detector.load_pickle(os.path.join(model_load_path, 'val_detection_score.pkl'))
        val_output = detector.load_pickle(os.path.join(model_load_path, 'val_output.pkl'))

        # Load inference results for test data
        test_detection_score_list = detector.load_pickle(os.path.join(model_load_path, 'test_detection_score.pkl'))
        test_output = detector.load_pickle(os.path.join(model_load_path, 'test_output.pkl'))

        # Evaluate the model
        threshold = detector.unsupervised_threshold(val_detection_score_list)

        groundtruth_labels, predicted_labels, total_delays = detector.evaluate_online(
            input_list=test_list,
            detection_score_list=test_detection_score_list,
            threshold=threshold,
        )

        results.append({
            'Seed': model_seed,
            'Fold': fold_idx,
            'F1': metrics.f1_score(groundtruth_labels, predicted_labels, zero_division=0.0),
            'Precision': metrics.precision_score(groundtruth_labels, predicted_labels, zero_division=0.0),
            'Recall': metrics.recall_score(groundtruth_labels, predicted_labels, zero_division=0.0),
            'Delay': np.mean(total_delays),
            'Threshold': threshold
        })

        f1_list = []
        reduced_test_detection_score = np.concatenate(test_detection_score_list).ravel()
        percentile_array = np.arange(0, 100.01, 0.01)
        for threshold_percentile in percentile_array:
            threshold_temp = np.percentile(reduced_test_detection_score, threshold_percentile)
            groundtruth_labels_temp, predicted_labels_temp, _ = detector.evaluate_online(
                input_list=test_list,
                detection_score_list=test_detection_score_list,
                threshold=threshold_temp,
            )
            f1_list.append(metrics.f1_score(groundtruth_labels_temp, predicted_labels_temp, zero_division=0.0))
        f1_list = np.vstack(f1_list)
        threshold_best = np.percentile(reduced_test_detection_score, percentile_array[np.argmax(f1_list)]).astype(float)

        groundtruth_labels_best, predicted_labels_best, total_delays_best = detector.evaluate_online(
            input_list=test_list,
            detection_score_list=test_detection_score_list,
            threshold=threshold_best,
        )

        results_best.append({
            'Seed': model_seed,
            'Fold': fold_idx,
            'F1': metrics.f1_score(groundtruth_labels_best, predicted_labels_best, zero_division=0.0),
            'Precision': metrics.precision_score(groundtruth_labels_best, predicted_labels_best, zero_division=0.0),
            'Recall': metrics.recall_score(groundtruth_labels_best, predicted_labels_best, zero_division=0.0),
            'Delay': np.mean(total_delays_best),
            'Threshold': threshold_best
        })

results = pd.DataFrame(results)
results_best = pd.DataFrame(results_best)

if not os.path.isfile(os.path.join(model_path, 'results.xlsx')):
    # Create and save a valid Excel file
    wb = openpyxl.Workbook()
    wb.save(os.path.join(model_path, 'results.xlsx'))

# Use a try-finally block to ensure proper handling
try:
    with pd.ExcelWriter(os.path.join(model_path, 'results.xlsx'), mode='a', if_sheet_exists='overlay') as writer:
        results.to_excel(writer, index=False, sheet_name=MODEL_NAME + '_' + AD_MODE)
        results_best.to_excel(writer, index=False, sheet_name=MODEL_NAME + '_' + AD_MODE + '_best')
finally:
    # Cleanup: Remove default 'Sheet' if it exists
    try:
        workbook = openpyxl.load_workbook(os.path.join(model_path, 'results.xlsx'))
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        workbook.save(os.path.join(model_path, 'results.xlsx'))
    except Exception as e:
        print(f"Error cleaning up sheets: {e}")
